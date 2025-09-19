import pytz
from django.contrib import admin, messages
from django.utils.html import format_html
from django.template.loader import render_to_string
from django.utils import timezone
from accounts.email_utils import send_email_async_task
from .models import (
    Presenter, Event, Presentation,
    SoloCompetition, GroupCompetition, CompetitionTeam, TeamMembership,
    TeamContent, ContentImage, ContentLike, ContentComment,
    PresentationEnrollment, SoloCompetitionRegistration, Post
)
from io import BytesIO
from openpyxl import Workbook
import datetime
import re
from django.http import HttpResponse

def _format_datetime(dt):
    iran_tz = pytz.timezone('Asia/Tehran')
    local_dt = timezone.localtime(dt, iran_tz)
    return local_dt.strftime('%Y/%m/%d %H:%M')

# admin actions
@admin.action(description='Export presentation participants to Excel (.xlsx)')
def export_presentation_enrollments(modeladmin, request, queryset):
    """
    Export enrollments of the selected Presentations into a single Excel workbook.
    Each selected presentation gets its own worksheet named <id>_<safe_title>.
    """
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    total_rows = 0
    for pres in queryset:
        safe_title = re.sub(r'[^0-9a-zA-Z\u0600-\u06FF ]+', '_', (pres.title or ""))
        sheet_name = f"{pres.id}_{safe_title}"[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Header row
        headers = [
            "user_id",
            "email",
            "full_name",
            "phone_number",  # new column
            "status",
            "order_item_id",
            "enrolled_at",
        ]
        ws.append(headers)

        enrollments = pres.enrollments.select_related('user', 'order_item').all()
        for en in enrollments:
            user = getattr(en, 'user', None)
            user_id = getattr(user, 'id', '') if user else ''
            email = getattr(user, 'email', '') if user else ''

            full_name = getattr(user, 'get_full_name', None)
            if callable(full_name):
                full_name = full_name()
            else:
                full_name = getattr(user, 'first_name', '') or getattr(user, 'username', '') or email

            phone_number = getattr(user, 'phone_number', '') if user else ''  # new line

            status = en.get_status_display() if hasattr(en, 'get_status_display') else en.status
            order_item_id = en.order_item.id if getattr(en, 'order_item', None) else ''
            enrolled_at = _format_datetime(en.enrolled_at) if en.enrolled_at else ''

            ws.append([user_id, email, full_name, phone_number, status, order_item_id, enrolled_at])
            total_rows += 1

        ws.append([])
        ws.append([f"Exported from admin at {datetime.datetime.utcnow().isoformat()}Z"])

    if total_rows == 0:
        modeladmin.message_user(request, "No enrollments found for the selected presentations.", level=messages.WARNING)
        return

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    ts = datetime.datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"presentation_enrollments_{ts}.xlsx"

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

@admin.action(description='Send reminder email to presentation participants')
def send_presentation_reminder(modeladmin, request, queryset):
    total = 0
    for pres in queryset:
        qs = pres.enrollments.filter(status=PresentationEnrollment.STATUS_COMPLETED_OR_FREE)
        emails = list(qs.values_list('user__email', flat=True))
        if not emails:
            continue
        html = render_to_string('reminder.html', {
            'object': pres,
            'object_datetime': _format_datetime(pres.start_time),
            'object_location': pres.online_link if pres.is_online else pres.location or '',
        })
        subject = f'یادآوری ارائه: {pres.title}'
        send_email_async_task(subject, emails, text_content='', html_content=html)
        total += len(emails)
    messages.success(request, f'{total} reminder emails sent.')


@admin.action(description='Send reminder email to solo competition participants')
def send_solo_competition_reminder(modeladmin, request, queryset):
    total = 0
    for comp in queryset:
        qs = comp.registrations.filter(status=SoloCompetitionRegistration.STATUS_COMPLETED_OR_FREE)
        emails = list(qs.values_list('user__email', flat=True))
        if not emails:
            continue
        html = render_to_string('reminder.html', {
            'object': comp,
            'object_datetime': _format_datetime(comp.start_datetime),
        })
        subject = f'یادآوری مسابقهٔ تکی: {comp.title}'
        send_email_async_task(subject, emails, text_content='', html_content=html)
        total += len(emails)
    messages.success(request, f'{total} emails sent.')

@admin.action(description='Export solo competition participants to Excel (.xlsx)')
def export_solo_competition_registrations(modeladmin, request, queryset):
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    total_rows = 0
    for comp in queryset:
        safe_title = re.sub(r'[^0-9a-zA-Z\u0600-\u06FF ]+', '_', comp.title or "")
        sheet_name = f"{comp.id}_{safe_title}"[:31]
        ws = wb.create_sheet(title=sheet_name)

        headers = [
            "user_id", "full_name", "email", "phone_number",
            "status", "order_item_id", "registered_at"
        ]
        ws.append(headers)

        registrations = comp.registrations.select_related('user', 'order_item').all()
        for reg in registrations:
            user = reg.user
            user_id = getattr(user, 'id', '') if user else ''
            email = getattr(user, 'email', '') if user else ''
            full_name = getattr(user, 'get_full_name', None)
            if callable(full_name):
                full_name = full_name()
            else:
                full_name = getattr(user, 'first_name', '') or getattr(user, 'username', '') or email

            phone_number = getattr(user, 'phone_number', '') if user else ''
            status = reg.get_status_display() if hasattr(reg, 'get_status_display') else reg.status
            order_item_id = reg.order_item.id if getattr(reg, 'order_item', None) else ''
            registered_at = _format_datetime(reg.registered_at) if reg.registered_at else ''

            ws.append([user_id, full_name, email, phone_number, status, order_item_id, registered_at])
            total_rows += 1

        ws.append([])
        ws.append([f"Exported from admin at {datetime.datetime.utcnow().isoformat()}Z"])

    if total_rows == 0:
        modeladmin.message_user(request, "No participants found for the selected solo competitions.", level=messages.WARNING)
        return

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    ts = datetime.datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"solo_competition_participants_{ts}.xlsx"

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@admin.action(description='Send reminder email to group competition participants')
def send_group_competition_reminder(modeladmin, request, queryset):
    total = 0
    for comp in queryset:
        teams = comp.teams.filter(status=CompetitionTeam.STATUS_ACTIVE)
        emails_set = set()
        for team in teams:
            if team.leader and team.leader.email:
                emails_set.add(team.leader.email)
            member_emails = team.memberships.values_list('user__email', flat=True)
            emails_set.update(member_emails)
        if not emails_set:
            continue
        html = render_to_string('reminder.html', {
            'object': comp,
            'object_datetime': _format_datetime(comp.start_datetime),
        })
        subject = f'یادآوری مسابقهٔ گروهی: {comp.title}'
        send_email_async_task(subject, list(emails_set), text_content='', html_content=html)
        total += len(emails_set)
    messages.success(request, f'{total} emails sent.')

@admin.action(description='Export group competition teams and members to Excel (.xlsx)')
def export_group_competition_teams(modeladmin, request, queryset):
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    total_rows = 0
    for comp in queryset:
        safe_title = re.sub(r'[^0-9a-zA-Z\u0600-\u06FF ]+', '_', comp.title or "")
        sheet_name = f"{comp.id}_{safe_title}"[:31]
        ws = wb.create_sheet(title=sheet_name)

        headers = [
            "team_id", "team_name",
            "leader_email",
            "member_name", "member_email", "member_phone",
            "status", "joined_at"
        ]
        ws.append(headers)

        teams = comp.teams.select_related('leader').prefetch_related('memberships__user').all()
        for team in teams:
            leader_email = team.leader.email if team.leader else ''
            status = team.status

            memberships = team.memberships.select_related('user').all()
            if memberships:
                for mem in memberships:
                    user = mem.user
                    member_name = user.get_full_name() if hasattr(user, 'get_full_name') else getattr(user, 'username', '')
                    member_email = user.email if user else ''
                    member_phone = getattr(user, 'phone_number', '') if user else ''
                    joined_at = _format_datetime(mem.joined_at) if mem.joined_at else ''
                    ws.append([team.id, team.name, leader_email,
                               member_name, member_email, member_phone,
                               status, joined_at])
                    total_rows += 1
            else:
                # Team has no members, still include leader email
                ws.append([team.id, team.name, leader_email,
                           '', '', '', status, ''])
                total_rows += 1

        ws.append([])
        ws.append([f"Exported from admin at {datetime.datetime.utcnow().isoformat()}Z"])

    if total_rows == 0:
        modeladmin.message_user(request, "No teams found for the selected group competitions.", level=messages.WARNING)
        return

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    ts = datetime.datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"group_competition_teams_{ts}.xlsx"

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response



@admin.register(Presenter)
class PresenterAdmin(admin.ModelAdmin):
    list_display = ("name", "email")
    search_fields = ("name", "email", "bio")


@admin.register(Presentation)
class PresentationAdmin(admin.ModelAdmin):
    list_display = ("title", "event", "type", "level", "start_time", "is_active", "is_paid")
    list_filter = ("is_active", "is_paid", "event", "type", "level")
    search_fields = ("title", "description", "event__title", "presenters__name")
    autocomplete_fields = ['event', 'presenters']
    filter_horizontal = ('presenters',)
    readonly_fields = ("poster_preview", "created_at")
    actions = [send_presentation_reminder, export_presentation_enrollments]
    fieldsets = (
        (None, {
            "fields": (
                "event", "title", "description", "presenters",
                "type", "level", "is_online", "location", "online_link",
                "start_time", "end_time", "is_active",
            )
        }),
        ("Payment", {"fields": ("is_paid", "price", "capacity")}),
        ("Media",   {"fields": ("poster", "poster_preview")}),
        ("Meta",    {"fields": ("created_at",)}),
    )

    @admin.display(description="Poster Preview")
    def poster_preview(self, obj):
        if obj.poster and hasattr(obj.poster, "url"):
            return format_html('<img src="{}" style="max-width:320px;height:auto;border-radius:6px;" />', obj.poster.url)
        return "No poster uploaded"


class PresentationInline(admin.TabularInline):
    model = Presentation
    extra = 0
    fields = ('title', 'type', 'start_time', 'end_time', 'is_active')
    show_change_link = True
    ordering = ('start_time',)


class SoloCompetitionInline(admin.TabularInline):
    model = SoloCompetition
    extra = 0
    fields = ('title', 'start_datetime', 'end_datetime', 'is_active')
    show_change_link = True
    ordering = ('start_datetime',)


class GroupCompetitionInline(admin.TabularInline):
    model = GroupCompetition
    extra = 0
    fields = ('title', 'start_datetime', 'end_datetime', 'is_active')
    show_change_link = True
    ordering = ('start_datetime',)


@admin.register(Event)
class EventAdmin(admin.ModelAdmin):
    list_display = ('id', 'title', 'start_date', 'end_date', 'is_active')
    search_fields = ('id', 'title', 'description')
    list_filter = ('is_active', 'start_date')
    inlines = [PresentationInline, SoloCompetitionInline, GroupCompetitionInline]
    readonly_fields = ('created_at',)

    def get_readonly_fields(self, request, obj=None):
        readonly_fields = list(self.readonly_fields)
        if obj:
            readonly_fields.append('id')
        return readonly_fields


@admin.register(SoloCompetition)
class SoloCompetitionAdmin(admin.ModelAdmin):
    list_display = ('title', 'event', 'start_datetime', 'is_active', 'is_paid')
    search_fields = ('title', 'description', 'event__title')
    list_filter = ('is_paid', 'is_active', 'event')
    autocomplete_fields = ['event']
    readonly_fields = ('created_at',)
    actions = [send_solo_competition_reminder, export_solo_competition_registrations]


@admin.register(GroupCompetition)
class GroupCompetitionAdmin(admin.ModelAdmin):
    list_display = ('title', 'event', 'start_datetime', 'is_active', 'requires_admin_approval')
    search_fields = ('title', 'description', 'event__title')
    list_filter = ('is_paid', 'is_active', 'requires_admin_approval', 'event')
    autocomplete_fields = ['event']
    readonly_fields = ('created_at',)
    actions = [send_group_competition_reminder, export_group_competition_teams]


class TeamMembershipInline(admin.TabularInline):
    model = TeamMembership
    extra = 1
    autocomplete_fields = ['user']
    readonly_fields = ('joined_at',)
    ordering = ('-joined_at',)


class ContentImageInline(admin.TabularInline):
    model = ContentImage
    extra = 1

class TeamContentInline(admin.StackedInline):
    model = TeamContent
    extra = 0
    readonly_fields = ('created_at',)
    inlines = [ContentImageInline]


@admin.register(CompetitionTeam)
class CompetitionTeamAdmin(admin.ModelAdmin):
    list_display = ('name', 'leader', 'group_competition', 'status', 'is_approved_by_admin')
    search_fields = ('name', 'leader__email', 'group_competition__title')
    list_filter = ('status', 'is_approved_by_admin', 'group_competition__event')
    autocomplete_fields = ['leader', 'group_competition']
    inlines = [TeamMembershipInline, TeamContentInline]
    readonly_fields = ('created_at',)
    list_select_related = ('leader', 'group_competition')

    @admin.action(description="Approve selected teams")
    def approve_teams(self, request, queryset):
        updated = 0
        for team in queryset:
            if team.needs_admin_approval() and team.status == CompetitionTeam.STATUS_PENDING_ADMIN_VERIFICATION:
                team.is_approved_by_admin = True
                team.status = (CompetitionTeam.STATUS_ACTIVE
                               if not team.group_competition.is_paid
                               or team.group_competition.price_per_member <= 0
                               else CompetitionTeam.STATUS_APPROVED_AWAITING_PAYMENT)
                team.save()
                updated += 1
        self.message_user(request, f"{updated} team(s) approved.")

    @admin.action(description="Reject selected teams")
    def reject_teams(self, request, queryset):
        updated = 0
        for team in queryset:
            if team.needs_admin_approval() and team.status == CompetitionTeam.STATUS_PENDING_ADMIN_VERIFICATION:
                team.is_approved_by_admin = False
                team.status = CompetitionTeam.STATUS_REJECTED_BY_ADMIN
                team.save()
                updated += 1
        self.message_user(request, f"{updated} team(s) rejected.")

    actions = ['approve_teams', 'reject_teams']


@admin.register(TeamMembership)
class TeamMembershipAdmin(admin.ModelAdmin):
    list_display = ('user', 'team', 'joined_at')
    search_fields = ('user__email', 'team__name')
    list_filter = ('team__group_competition',)
    autocomplete_fields = ['user', 'team']
    readonly_fields = ('joined_at',)


@admin.register(PresentationEnrollment)
class PresentationEnrollmentAdmin(admin.ModelAdmin):
    list_display = ('user', 'presentation', 'status', 'enrolled_at')
    search_fields = ('user__email', 'presentation__title')
    list_filter = ('status', 'presentation__event')
    autocomplete_fields = ['user', 'presentation', 'order_item']
    readonly_fields = ('enrolled_at',)
    list_select_related = ('user', 'presentation')


@admin.register(SoloCompetitionRegistration)
class SoloCompetitionRegistrationAdmin(admin.ModelAdmin):
    list_display = ('user', 'solo_competition', 'status', 'registered_at')
    search_fields = ('user__email', 'solo_competition__title')
    list_filter = ('status', 'solo_competition__event')
    autocomplete_fields = ['user', 'solo_competition', 'order_item']
    readonly_fields = ('registered_at',)
    list_select_related = ('user', 'solo_competition')


@admin.register(TeamContent)
class TeamContentAdmin(admin.ModelAdmin):
    list_display = ('team', 'file_link', 'created_at')
    search_fields = ('team__name',)
    autocomplete_fields = ['team']
    inlines = [ContentImageInline]

@admin.register(ContentLike)
class ContentLikeAdmin(admin.ModelAdmin):
    list_display = ('user', 'team_content', 'created_at')
    search_fields = ('user__email', 'team_content__team__name')
    autocomplete_fields = ['user', 'team_content']


@admin.register(ContentComment)
class ContentCommentAdmin(admin.ModelAdmin):
    list_display = ('user', 'text_snippet', 'created_at')
    search_fields = ('user__email', 'text')
    autocomplete_fields = ['user', 'team_content']

    @admin.display(description="Comment")
    def text_snippet(self, obj):
        return (obj.text[:75] + '...') if len(obj.text) > 75 else obj.text


@admin.register(Post)
class PostAdmin(admin.ModelAdmin):
    list_display = ("title", "is_active", "published_at")
    list_filter = ("is_active",)
    search_fields = ("title", "excerpt", "body_markdown")
    ordering = ("-published_at",)
